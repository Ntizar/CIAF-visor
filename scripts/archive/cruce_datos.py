#!/usr/bin/env python3
"""
CIAF Data Cross-Reference Tool
Cruza datos del Excel (fuente de verdad) con los JSONs existentes.
Corrige severidad, tipología y geolocalización.

Autor: David Antizar (ejecutado por Mastermind)
"""

import os
import sys
import json
import re
import time
import math
from collections import Counter, defaultdict
from datetime import datetime

# ─── CONFIGURACIÓN ───────────────────────────────────────────────
EXCEL_PATH = "/persist/hermes-home/webui/attachments/591876dde7a5/260218_Base_Datos_CIAF_1.xlsx"
JSON_DIR = "/root/workspace/ciaf-data/data/individual"
OUTPUT_DIR = "/root/workspace/ciaf-data/data/individual_corrected"
REPORT_PATH = "/root/workspace/CIAF-visor/cruce-datos-reporte.md"

# ─── TAXONOMÍA RD 929/2022 ──────────────────────────────────────
# Mapeo de tipo_suceso del Excel → categoría normalizada
TIPO_NORMALIZED = {
    # Accidentes muy graves
    "descarrilamiento": "accidente",
    "descarrilamiento de tren": "accidente",
    "accidente grave por descarrilamiento": "accidente",
    "descarrilamiento por escape de material": "accidente",
    # Colisiones
    "colisión": "accidente",
    "colisión de trenes": "accidente",
    "colisión entre trenes": "accidente",
    "colisión de tren": "accidente",
    "colisión frontal": "accidente",
    "colisión frontal de trenes": "accidente",
    "colisión lateral": "accidente",
    "colisión por alcance": "accidente",
    "colisión entre tren y vehículo de carretera": "accidente",
    "colisión con roca": "accidente",
    "colisión con desprendimiento de rocas": "accidente",
    "colisión con obstáculos en la vía": "accidente",
    "colisión con obstáculo en la vía": "accidente",
    "colisión y descarrilamiento": "accidente",
    "colisión de tren con señal y maquinaria en vía": "accidente",
    # Arrollamientos
    "arrollamiento": "accidente",
    "arrollamiento de persona": "accidente",
    "arrollamiento de personas": "accidente",
    "arrollamiento de peatón": "accidente",
    "arrollamiento de vehículo por tren": "accidente",
    "arrollamiento de un vehículo por tren": "accidente",
    "arrollamiento de ciclista": "accidente",
    "arrollamiento de motocicleta por tren": "accidente",
    "arrollamiento de obstáculo": "accidente",
    "arrollamiento de obstáculo y descarrilamiento": "accidente",
    "arrollamiento de barra de carril": "accidente",
    # Paso a nivel
    "accidente en paso a nivel": "accidente",
    # Incendios
    "incendio de material rodante": "accidente",
    "incendio en locomotora": "accidente",
    "incendio en tren": "accidente",
    # Genéricos
    "accidente ferroviario": "accidente",
    # Conatos (casi-accidente)
    "conato de colisión": "incidente",
    "conato de colisión entre trenes": "incidente",
    "conato de colisión por alcance": "incidente",
    "conato de colisión por rebase de señal": "incidente",
    "conato de colisión y descarrilamiento": "incidente",
    "conato de incendio": "incidente",
    # Incidentes operacionales
    "incidente ferroviario": "incidente",
    "incidente operacional": "incidente",
    "incidente operacional con riesgo de colisión": "incidente",
    "incidente operacional - retroceso no autorizado": "incidente",
    # Rebase de señal
    "rebase de señal": "incidente",
    "rebase indebido de señal": "incidente",
    "rebase de señal en rojo": "incidente",
    "rebase de señal de parada": "incidente",
    "rebase de señal y deriva de locomotora": "incidente",
    # Escape de material
    "escape de material": "incidente",
    "escape de tren a la deriva": "incidente",
    # Fallos
    "fallo de señalización": "incidente",
    "fallo en las instalaciones de seguridad": "incidente",
    "fallo de cargamento": "incidente",
    "rotura de eje en tren de viajeros": "incidente",
}

# Subcategoría detallada (para campo tipo_suceso_normalizado)
TIPO_DETALLE = {
    "descarrilamiento": "descarrilamiento",
    "descarrilamiento de tren": "descarrilamiento",
    "accidente grave por descarrilamiento": "descarrilamiento",
    "descarrilamiento por escape de material": "descarrilamiento",
    "colisión": "colision",
    "colisión de trenes": "colision_trenes",
    "colisión entre trenes": "colision_trenes",
    "colisión de tren": "colision_trenes",
    "colisión frontal": "colision_frontal",
    "colisión frontal de trenes": "colision_frontal",
    "colisión lateral": "colision_lateral",
    "colisión por alcance": "colision_alcance",
    "colisión entre tren y vehículo de carretera": "colision_vehiculo",
    "colisión con roca": "colision_rocas",
    "colisión con desprendimiento de rocas": "colision_rocas",
    "colisión con obstáculos en la vía": "colision_obstaculo",
    "colisión con obstáculo en la vía": "colision_obstaculo",
    "colisión y descarrilamiento": "colision_descarrilamiento",
    "colisión de tren con señal y maquinaria en vía": "colision_infraestructura",
    "arrollamiento": "arrollamiento",
    "arrollamiento de persona": "arrollamiento_persona",
    "arrollamiento de personas": "arrollamiento_persona",
    "arrollamiento de peatón": "arrollamiento_persona",
    "arrollamiento de vehículo por tren": "arrollamiento_vehiculo",
    "arrollamiento de un vehículo por tren": "arrollamiento_vehiculo",
    "arrollamiento de ciclista": "arrollamiento_ciclista",
    "arrollamiento de motocicleta por tren": "arrollamiento_ciclista",
    "arrollamiento de obstáculo": "arrollamiento_obstaculo",
    "arrollamiento de obstáculo y descarrilamiento": "arrollamiento_obstaculo",
    "arrollamiento de barra de carril": "arrollamiento_obstaculo",
    "accidente en paso a nivel": "paso_nivel",
    "incendio de material rodante": "incendio",
    "incendio en locomotora": "incendio",
    "incendio en tren": "incendio",
    "accidente ferroviario": "otro_accidente",
    "conato de colisión": "conato_colision",
    "conato de colisión entre trenes": "conato_colision",
    "conato de colisión por alcance": "conato_colision",
    "conato de colisión por rebase de señal": "conato_colision",
    "conato de colisión y descarrilamiento": "conato_colision",
    "conato de incendio": "conato_incendio",
    "incidente ferroviario": "otro_incidente",
    "incidente operacional": "incidente_operacional",
    "incidente operacional con riesgo de colisión": "incidente_operacional",
    "incidente operacional - retroceso no autorizado": "incidente_operacional",
    "rebase de señal": "rebase_senal",
    "rebase indebido de señal": "rebase_senal",
    "rebase de señal en rojo": "rebase_senal",
    "rebase de señal de parada": "rebase_senal",
    "rebase de señal y deriva de locomotora": "rebase_senal",
    "escape de material": "escape_material",
    "escape de tren a la deriva": "escape_material",
    "fallo de señalización": "fallo_senalizacion",
    "fallo en las instalaciones de seguridad": "fallo_seguridad",
    "fallo de cargamento": "fallo_cargamento",
    "rotura de eje en tren de viajeros": "rotura_eje",
}


# ─── FUNCIONES DE MATCHING ──────────────────────────────────────

def normalize_exp(exp_str):
    """Normalizar número de expediente: '0001/2008' → '1/2008'"""
    if not exp_str:
        return None
    exp = str(exp_str).strip()
    m = re.match(r'^0*(\d+)/(\d{4})$', exp)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    return exp


def extract_exp_from_json(data, filename):
    """Extraer número de expediente de un JSON CIAF."""
    # 1. Buscar en el título
    titulo = data.get('titulo', data.get('title', ''))
    patterns = [
        r'N[º°]?\s*(\d{3,4}/\d{4})',
        r'IF\s+(\d+/\d{4})',
        r'expediente\s+n[º°]?\s*(\d+/\d{4})',
        r'(\d{3,4}/\d{4})',
    ]
    for pat in patterns:
        m = re.search(pat, titulo, re.IGNORECASE)
        if m:
            return normalize_exp(m.group(1))
    
    # 2. Buscar en el ID
    jid = data.get('id', '')
    # IDs como "IF-0002-2010" o "0064-2010"
    m = re.search(r'(\d{3,4})-(\d{4})', jid)
    if m:
        return normalize_exp(f"{m.group(1)}/{m.group(2)}")
    
    # 3. Del nombre de archivo
    fname = filename.replace('.json', '').replace('CIAF', '')
    year = data.get('año', '')
    numbers = re.findall(r'(\d{3,4})', fname)
    for n in numbers:
        if int(n) < 500 and year:
            return normalize_exp(f"{n}/{year}")
    
    return None


# ─── FUNCIONES DE SEVERIDAD ─────────────────────────────────────

def compute_severity_from_excel(row):
    """Calcular severidad del Excel según RD 929/2022."""
    muertos = row.get('muertos', 0) or 0
    hg = row.get('heridos_graves', 0) or 0
    
    if muertos > 0:
        return "muy grave"
    elif hg > 0:
        return "grave"
    else:
        return "menor"


def compute_severity_from_json(data):
    """Calcular severidad de un JSON existente."""
    fallecidos = data.get('victimas_fallecidos', 0) or 0
    graves = data.get('victimas_graves', 0) or 0
    
    if fallecidos > 0:
        return "muy grave"
    elif graves > 0:
        return "grave"
    else:
        return "menor"


# ─── FUNCIONES DE PK ────────────────────────────────────────────

def parse_pk(pk_str):
    """
    Parsear PK en formato estándar.
    Acepta: '429,825', '415+648', 'P.K. 62+902', 'PK 166+800'
    Retorna: (pk_km, pk_m) o None
    """
    if not pk_str:
        return None
    
    pk = str(pk_str).strip()
    
    # Formato europeo: 429,825 (coma como decimal)
    m = re.match(r'^(\d+)[,.](\d+)$', pk)
    if m:
        return float(f"{m.group(1)}.{m.group(2)}")
    
    # Formato estándar: 415+648 o P.K. 62+902
    m = re.search(r'(\d+)\+(\d+)', pk)
    if m:
        km = int(m.group(1))
        m_part = int(m.group(2))
        return km + m_part / 1000.0
    
    # Solo número
    m = re.match(r'^(\d+\.?\d*)$', pk)
    if m:
        return float(m.group(1))
    
    return None


def normalize_pk(pk_str):
    """Normalizar PK a formato 'NNN+NNN'."""
    val = parse_pk(pk_str)
    if val is None:
        return pk_str  # devolver original si no se puede parsear
    
    km = int(val)
    m = int(round((val - km) * 1000))
    return f"{km}+{m:03d}"


# ─── FUNCIONES DE TIPOLOGÍA ────────────────────────────────────

def normalize_tipo(tipo_suceso):
    """Normalizar tipo_suceso del Excel a categoría estándar."""
    if not tipo_suceso:
        return "desconocido", "desconocido"
    
    tipo_lower = tipo_suceso.lower().strip()
    
    categoria = TIPO_NORMALIZED.get(tipo_lower, "otro")
    detalle = TIPO_DETALLE.get(tipo_lower, "otro")
    
    return categoria, detalle


# ─── FUNCIONES DE GEOCODIFICACIÓN ──────────────────────────────

def geocode_by_line_pk(linea, pk_val, tramo=""):
    """
    Geocodificar usando línea + PK.
    Por ahora retorna None (necesita API de ADIF).
    En implementación futura: interpolar geometría WFS de Tramificación.
    """
    # TODO: Implementar con ADIF WFS Tramificación
    # 1. Buscar tramo por nombre de línea en WFS
    # 2. Obtener geometría del tramo
    # 3. Interpolar posición a lo largo del PK
    # 4. Retornar (lat, lng)
    return None, None


# ─── SCRIPT PRINCIPAL ──────────────────────────────────────────

def main():
    print("=" * 60)
    print("  CIAF Data Cross-Reference Tool")
    print("  Cruza Excel ↔ JSONs, corrige severidad, tipología y PK")
    print("=" * 60)
    
    # ─── FASE 1: Cargar Excel ───
    print("\n📊 FASE 1: Cargando Excel...")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        ws = wb.active
        headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        
        excel_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                excel_rows.append(dict(zip(headers, row)))
        
        # Indexar por expediente único (primera ocurrencia)
        excel_by_exp = {}
        excel_all_rows = defaultdict(list)
        for r in excel_rows:
            exp = normalize_exp(r.get('n_exp', ''))
            if exp:
                excel_all_rows[exp].append(r)
                if exp not in excel_by_exp:
                    excel_by_exp[exp] = r
        
        print(f"  ✅ {len(excel_rows)} filas totales")
        print(f"  ✅ {len(excel_by_exp)} expedientes únicos")
        print(f"  ✅ {len(excel_all_rows)} con múltiples filas (recomendaciones)")
        wb.close()
    except Exception as e:
        print(f"  ❌ Error cargando Excel: {e}")
        return
    
    # ─── FASE 2: Cargar JSONs ───
    print("\n📁 FASE 2: Cargando JSONs...")
    json_files = sorted([f for f in os.listdir(JSON_DIR) if f.endswith('.json')])
    json_by_exp = {}
    
    for f in json_files:
        with open(os.path.join(JSON_DIR, f), 'r') as fh:
            data = json.load(fh)
        exp = extract_exp_from_json(data, f)
        if exp:
            json_by_exp[exp] = {'file': f, 'data': data}
    
    print(f"  ✅ {len(json_files)} JSONs cargados")
    print(f"  ✅ {len(json_by_exp)} con expediente extraído")
    
    # ─── FASE 3: Matching ───
    print("\n🔗 FASE 3: Emparejando registros...")
    matched = []
    unmatched_excel = []
    unmatched_json = []
    
    for exp, excel_data in excel_by_exp.items():
        if exp in json_by_exp:
            matched.append({
                'exp': exp,
                'excel': excel_data,
                'json': json_by_exp[exp]['data'],
                'json_file': json_by_exp[exp]['file'],
                'excel_rows': excel_all_rows[exp]
            })
        else:
            unmatched_excel.append(exp)
    
    for exp, jdata in json_by_exp.items():
        if exp not in excel_by_exp:
            unmatched_json.append({'exp': exp, 'file': jdata['file'], 'data': jdata['data']})
    
    print(f"  ✅ Emparejados: {len(matched)}")
    print(f"  ⚠️  Excel sin JSON: {len(unmatched_excel)}")
    print(f"  ⚠️  JSON sin Excel: {len(unmatched_json)}")
    
    # ─── FASE 4: Correcciones ───
    print("\n🔧 FASE 4: Aplicando correcciones...")
    
    stats = {
        'severidad_corregida': 0,
        'tipo_corregido': 0,
        'pk_normalizado': 0,
        'campos_anadidos': 0,
        'total_match': len(matched),
    }
    
    corrected = []
    
    for item in matched:
        exp = item['exp']
        excel = item['excel']
        json_data = item['json'].copy()  # no modificar original aún
        changes = []
        
        # 4.1 Corregir severidad
        excel_severity = compute_severity_from_excel(excel)
        json_severity = json_data.get('gravedad', '')
        
        if excel_severity != json_severity:
            json_data['gravedad'] = excel_severity
            json_data['gravedad_anterior'] = json_severity
            changes.append(f"severidad: {json_severity} → {excel_severity}")
            stats['severidad_corregida'] += 1
        
        # 4.2 Corregir tipología
        excel_tipo = excel.get('tipo_suceso', '')
        if excel_tipo:
            cat, detalle = normalize_tipo(excel_tipo)
            json_data['tipo'] = cat
            json_data['tipo_suceso'] = excel_tipo
            json_data['tipo_suceso_normalizado'] = detalle
            if cat != json_data.get('tipo', ''):
                changes.append(f"tipo: {json_data.get('tipo', '?')} → {cat} ({detalle})")
                stats['tipo_corregido'] += 1
            else:
                # Añadir campos detallados aunque el tipo no cambie
                stats['campos_anadidos'] += 1
        
        # 4.3 Normalizar PK
        excel_pk = excel.get('pk', '')
        if excel_pk:
            pk_normalized = normalize_pk(excel_pk)
            if pk_normalized != json_data.get('pk', ''):
                json_data['pk_original'] = json_data.get('pk', '')
                json_data['pk'] = pk_normalized
                stats['pk_normalizado'] += 1
            
            # Añadir PK numérico para cálculos
            pk_val = parse_pk(excel_pk)
            if pk_val:
                json_data['pk数值'] = round(pk_val, 3)
            
            # Añadir línea del Excel si el JSON no la tiene
            excel_linea = excel.get('linea', '')
            if excel_linea and not json_data.get('tramo'):
                json_data['tramo'] = excel_linea
                stats['campos_anadidos'] += 1
        
        # 4.4 Añadir información del Excel que falta
        # Descripción detallada
        excel_desc = excel.get('descripcion', '')
        if excel_desc and not json_data.get('resumen'):
            json_data['resumen'] = excel_desc
            stats['campos_anadidos'] += 1
        
        # Recomendaciones del Excel (todas las filas)
        if item['excel_rows']:
            recs = []
            for row in item['excel_rows']:
                rec_text = row.get('recomendacion_texto', '')
                rec_dest = row.get('recomendacion_destinatario', '')
                if rec_text and rec_text != 'N/A':
                    recs.append({
                        'texto': rec_text,
                        'destinatario': rec_dest or 'N/A'
                    })
            if recs:
                json_data['recomendaciones_excel'] = recs
        
        # Causa directa
        excel_causa = excel.get('causa_directa', '')
        if excel_causa:
            json_data['causa_directa'] = excel_causa
        
        # Factores contribuyentes
        excel_factores = excel.get('factores_contribuyentes', '')
        if excel_factores:
            json_data['factores_contribuyentes'] = excel_factores
        
        # Infraestructura
        excel_infra = excel.get('infraestructura', '')
        if excel_infra:
            json_data['infraestructura'] = excel_infra
        
        # Tiempo de afección
        excel_tiempo = excel.get('tiempo_afeccion', '')
        if excel_tiempo:
            json_data['tiempo_afeccion'] = excel_tiempo
        
        json_data['_cross_ref'] = {
            'excel_exp': exp,
            'excel_file': excel.get('archivo_fuente', ''),
            'corrected_at': datetime.now().isoformat(),
            'changes': changes
        }
        
        corrected.append({
            'exp': exp,
            'data': json_data,
            'file': item['json_file'],
            'changes': changes
        })
    
    # ─── FASE 5: Guardar resultados ───
    print("\n💾 FASE 5: Guardando resultados...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    for item in corrected:
        out_path = os.path.join(OUTPUT_DIR, item['file'])
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(item['data'], f, ensure_ascii=False, indent=2)
    
    print(f"  ✅ {len(corrected)} JSONs guardados en {OUTPUT_DIR}")
    
    # ─── FASE 6: Informe ───
    print("\n📝 FASE 6: Generando informe...")
    
    report = f"""# 🔗 Informe de Cruce de Datos CIAF
**Fecha:** {datetime.now().strftime('%Y-%m-%d %H:%M')}
**Excel:** {os.path.basename(EXCEL_PATH)}
**JSONs:** {len(json_files)} archivos en {JSON_DIR}

---

## 📊 Resumen del Cruce

| Métrica | Valor |
|---------|-------|
| Expedientes Excel | {len(excel_by_exp)} |
| JSONs cargados | {len(json_files)} |
| **Emparejados** | **{stats['total_match']}** |
| Excel sin JSON | {len(unmatched_excel)} |
| JSON sin Excel | {len(unmatched_json)} |

## 🔧 Correcciones Aplicadas

| Corrección | Cantidad |
|------------|----------|
| Severidad corregida | {stats['severidad_corregida']} |
| Tipología corregida/anadida | {stats['tipo_corregido']} |
| PK normalizados | {stats['pk_normalizado']} |
| Campos adicionados | {stats['campos_anadidos']} |

## 📋 Detalle de Cambios por Informe

"""
    
    # Mostrar solo informes con cambios significativos
    changes_count = Counter()
    for item in corrected:
        for c in item['changes']:
            field = c.split(':')[0]
            changes_count[field] += 1
    
    report += "### Cambios por campo:\n"
    for field, count in changes_count.most_common():
        report += f"- **{field}:** {count} informes\n"
    
    report += f"\n### Muestra de cambios (primeros 20):\n\n"
    shown = 0
    for item in corrected:
        if item['changes'] and shown < 20:
            report += f"**{item['exp']}:**\n"
            for c in item['changes']:
                report += f"  - {c}\n"
            report += "\n"
            shown += 1
    
    # Informes sin JSON
    if unmatched_excel:
        report += f"\n## ⚠️ Expedientes sin JSON ({len(unmatched_excel)})\n\n"
        for exp in sorted(unmatched_excel):
            report += f"- {exp}\n"
    
    # JSONs sin Excel
    if unmatched_json:
        report += f"\n## ⚠️ JSONs sin Excel ({len(unmatched_json)})\n\n"
        for item in unmatched_json:
            report += f"- {item['file']} (id: {item['data'].get('id', '?')})\n"
    
    report += f"""
## 📊 Distribución de Severidad Corregida

"""
    
    # Contar severidad corregida
    sev_count = Counter()
    for item in corrected:
        sev_count[item['data'].get('gravedad', 'N/A')] += 1
    
    for sev, count in sev_count.most_common():
        report += f"- **{sev}:** {count} informes\n"
    
    report += f"""

## 📊 Distribución de Tipología Normalizada

"""
    
    tipo_count = Counter()
    for item in corrected:
        tipo_count[item['data'].get('tipo_suceso_normalizado', 'N/A')] += 1
    
    for tipo, count in tipo_count.most_common():
        report += f"- **{tipo}:** {count} informes\n"
    
    report += f"""

## 🗂️ Estructura de Salida

```
{OUTPUT_DIR}/
├── {len(corrected)} JSONs corregidos
└── (mismos nombres que originales)
```

## ⚠️ Próximos Pasos

1. **Geocodificación por PK:** Implementar interpolación usando WFS Tramificación ADIF
2. **Re-parseo de PDFs faltantes:** Para los {len(unmatched_excel)} informes sin JSON
3. **Validación cruzada:** Verificar que las víctimas del Excel coinciden con las del JSON

---
*Generado automáticamente por CIAF Cross-Reference Tool*
"""
    
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"  ✅ Informe guardado en {REPORT_PATH}")
    
    # ─── RESUMEN FINAL ───
    print("\n" + "=" * 60)
    print("  RESUMEN FINAL")
    print("=" * 60)
    print(f"  ✅ {stats['total_match']} informes cruzados")
    print(f"  🔧 {stats['severidad_corregida']} severidades corregidas")
    print(f"  🔧 {stats['tipo_corregido']} tipologías corregidas")
    print(f"  🔧 {stats['pk_normalizado']} PKs normalizados")
    print(f"  📁 JSONs corregidos en: {OUTPUT_DIR}")
    print(f"  📝 Informe: {REPORT_PATH}")
    print(f"  ⚠️  {len(unmatched_excel)} informes sin JSON")
    print(f"  ⚠️  {len(unmatched_json)} JSONs sin Excel")
    print("=" * 60)


if __name__ == '__main__':
    main()
