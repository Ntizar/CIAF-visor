#!/usr/bin/env python3
"""
CIAF Visor Fixer
Corrige gravedad, tipología, nombres de estación y geolocalización
en los archivos YYYY.json del visor.
"""

import json, os, re, time
from datetime import datetime
from collections import Counter

# ─── PATHS ──────────────────────────────────────────────────────
REPORT_DIR = "/root/workspace/CIAF-visor/data/reports"
STATION_COORDS_PATH = "/root/workspace/CIAF-visor/data/station-coords.json"
INDIVIDUAL_DIR = "/root/workspace/ciaf-data/data/individual"
EXCEL_PATH = "/persist/hermes-home/webui/attachments/591876dde7a5/260218_Base_Datos_CIAF_1.xlsx"

# ─── GRAVEDAD MAP ───────────────────────────────────────────────
GRAVEDAD_MAP = {
    "fatal": "muy grave",
    "leve": "menor",
    "grave": "grave",
    "muy grave": "muy grave",
    "menor": "menor",
}

# ─── TIPOLOGY NORMALIZATION ────────────────────────────────────
TIPO_MAP = {
    "descarrilamiento": ("accidente", "descarrilamiento"),
    "descarrilamiento de tren": ("accidente", "descarrilamiento"),
    "colisión": ("accidente", "colision"),
    "colisión de trenes": ("accidente", "colision_trenes"),
    "colisión entre trenes": ("accidente", "colision_trenes"),
    "colisión frontal": ("accidente", "colision_frontal"),
    "colisión frontal de trenes": ("accidente", "colision_frontal"),
    "colisión lateral": ("accidente", "colision_lateral"),
    "colisión por alcance": ("accidente", "colision_alcance"),
    "colisión entre tren y vehículo de carretera": ("accidente", "colision_vehiculo"),
    "colisión con roca": ("accidente", "colision_rocas"),
    "colisión con desprendimiento de rocas": ("accidente", "colision_rocas"),
    "colisión con obstáculos en la vía": ("accidente", "colision_obstaculo"),
    "colisión con obstáculo en la vía": ("accidente", "colision_obstaculo"),
    "colisión y descarrilamiento": ("accidente", "colision_descarrilamiento"),
    "colisión de tren con señal y maquinaria en vía": ("accidente", "colision_infraestructura"),
    "arrollamiento": ("accidente", "arrollamiento"),
    "arrollamiento de persona": ("accidente", "arrollamiento_persona"),
    "arrollamiento de personas": ("accidente", "arrollamiento_persona"),
    "arrollamiento de peatón": ("accidente", "arrollamiento_persona"),
    "arrollamiento de vehículo por tren": ("accidente", "arrollamiento_vehiculo"),
    "arrollamiento de un vehículo por tren": ("accidente", "arrollamiento_vehiculo"),
    "arrollamiento de ciclista": ("accidente", "arrollamiento_ciclista"),
    "arrollamiento de motocicleta por tren": ("accidente", "arrollamiento_ciclista"),
    "arrollamiento de obstáculo": ("accidente", "arrollamiento_obstaculo"),
    "arrollamiento de obstáculo y descarrilamiento": ("accidente", "arrollamiento_obstaculo"),
    "arrollamiento de barra de carril": ("accidente", "arrollamiento_obstaculo"),
    "accidente en paso a nivel": ("accidente", "paso_nivel"),
    "incendio de material rodante": ("accidente", "incendio"),
    "incendio en locomotora": ("accidente", "incendio"),
    "incendio en tren": ("accidente", "incendio"),
    "accidente ferroviario": ("accidente", "otro_accidente"),
    "conato de colisión": ("incidente", "conato_colision"),
    "conato de colisión entre trenes": ("incidente", "conato_colision"),
    "conato de colisión por alcance": ("incidente", "conato_colision"),
    "conato de colisión por rebase de señal": ("incidente", "conato_colision"),
    "conato de colisión y descarrilamiento": ("incidente", "conato_colision"),
    "conato de incendio": ("incidente", "conato_incendio"),
    "incidente ferroviario": ("incidente", "otro_incidente"),
    "incidente operacional": ("incidente", "incidente_operacional"),
    "incidente operacional con riesgo de colisión": ("incidente", "incidente_operacional"),
    "incidente operacional - retroceso no autorizado": ("incidente", "incidente_operacional"),
    "rebase de señal": ("incidente", "rebase_senal"),
    "rebase indebido de señal": ("incidente", "rebase_senal"),
    "rebase de señal en rojo": ("incidente", "rebase_senal"),
    "rebase de señal de parada": ("incidente", "rebase_senal"),
    "rebase de señal y deriva de locomotora": ("incidente", "rebase_senal"),
    "escape de material": ("incidente", "escape_material"),
    "escape de tren a la deriva": ("incidente", "escape_material"),
    "fallo de señalización": ("incidente", "fallo_senalizacion"),
    "fallo en las instalaciones de seguridad": ("incidente", "fallo_seguridad"),
    "fallo de cargamento": ("incidente", "fallo_cargamento"),
    "rotura de eje en tren de viajeros": ("incidente", "rotura_eje"),
    "avería": ("incidente", "averia"),
}

# ─── STATION NAME CLEANING ─────────────────────────────────────

# Province names that shouldn't be station names
PROVINCE_NAMES = {
    'álava', 'albacete', 'alicante', 'almería', 'asturias', 'ávila',
    'badajoz', 'barcelona', 'bizkaia', 'burgos', 'cáceres', 'cádiz',
    'cantabria', 'castellón', 'ciudad real', 'córdoba', 'a coruña',
    'cuenca', 'gipuzkoa', 'girona', 'granada', 'guadalajara',
    'guipúzcoa', 'huelva', 'huesca', 'illes balears', 'islas baleares',
    'jaén', 'león', 'lleida', 'lugo', 'madrid', 'málaga', 'murcia',
    'navarra', 'orense', 'ourense', 'palencia', 'pontevedra',
    'la rioja', 'salamanca', 'santa cruz de tenerife', 'segovia',
    'sevilla', 'soria', 'tarragona', 'teruel', 'toledo', 'valencia',
    'valladolid', 'vizcaya', 'zamora', 'zaragoza',
}

# PK patterns to remove from station names
PK_PATTERNS = [
    r'P\.?K\.?\s*\d+[+,]\d+',
    r'p\.k\.\s*\d+[+,]\d+',
    r'PK\s*\d+[+,]\d+',
    r'pk\s*\d+[+,]\d+',
]

def clean_station_name(name):
    """Clean station name: remove periods, fix case, remove PK patterns."""
    if not name:
        return name
    
    cleaned = name.strip()
    
    # Remove trailing periods
    cleaned = cleaned.rstrip('.')
    
    # Remove PK patterns
    for pat in PK_PATTERNS:
        cleaned = re.sub(pat, '', cleaned, flags=re.IGNORECASE).strip()
    
    # Remove "Paso a nivel" prefix if it's just a PK reference
    if cleaned.lower().startswith('paso a nivel'):
        cleaned = re.sub(r'^Paso a nivel\s*', '', cleaned, flags=re.IGNORECASE).strip()
    
    # If empty after cleaning, return original
    if not cleaned:
        return name
    
    # Title case for ALL UPPERCASE names (but keep known abbreviations)
    if cleaned.isupper():
        cleaned = cleaned.title()
        # Fix common title-case issues
        cleaned = cleaned.replace('De ', 'de ').replace('Del ', 'del ')
        cleaned = cleaned.replace('La ', 'la ').replace('El ', 'el ')
        cleaned = cleaned.replace('Los ', 'los ').replace('Las ', 'las ')
        cleaned = cleaned.replace('A ', 'a ').replace('Y ', 'y ')
        # Fix province names back to proper case
        if cleaned.lower() in PROVINCE_NAMES:
            cleaned = cleaned  # keep as-is, will be flagged
    
    # Fix specific known issues
    fixes = {
        'Maçanet - Massanes.': 'Maçanet-Massanes',
        'Maçanet Massanes': 'Maçanet-Massanes',
        'Pinar De Las Rozas': 'Pinar de las Rozas',
        'Pinar de las Rozas': 'Pinar de las Rozas',
        'El Entrego': 'El Entrego',
        'Santa María De Huerta': 'Santa María de Huerta',
        'Santa María de Huerta': 'Santa María de Huerta',
        'Villalba De Guadarrama': 'Villalba de Guadarrama',
        'Villalba de Guadarrama': 'Villalba de Guadarrama',
        'Moncófar': 'Moncofa',
    }
    
    if cleaned in fixes:
        cleaned = fixes[cleaned]
    
    return cleaned


# ─── GEOCODING ──────────────────────────────────────────────────

def load_station_coords():
    """Load station coordinates database."""
    if os.path.exists(STATION_COORDS_PATH):
        with open(STATION_COORDS_PATH) as f:
            return json.load(f)
    return {}


def geocode_station(name, station_db):
    """Look up station coordinates from DB."""
    if not name:
        return None, None
    
    # Try exact match
    if name in station_db:
        coords = station_db[name]
        return coords.get('lat'), coords.get('lng')
    
    # Try case-insensitive
    name_lower = name.lower()
    for k, v in station_db.items():
        if k.lower() == name_lower:
            return v.get('lat'), v.get('lng')
    
    # Try partial match
    for k, v in station_db.items():
        if name_lower in k.lower() or k.lower() in name_lower:
            return v.get('lat'), v.get('lng')
    
    return None, None


# ─── LOAD EXCEL DATA ────────────────────────────────────────────

def load_excel_data():
    """Load Excel data for cross-reference."""
    import openpyxl
    
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    excel_by_exp = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            d = dict(zip(headers, row))
            exp = str(d.get('n_exp', '')).strip()
            m = re.match(r'^0*(\d+)/(\d{4})$', exp)
            if m:
                exp = f"{m.group(1)}/{m.group(2)}"
            if exp not in excel_by_exp:
                excel_by_exp[exp] = d
    
    wb.close()
    return excel_by_exp


# ─── BUILD INDIVIDUAL JSON INDEX ───────────────────────────────

def build_individual_index():
    """Build index of individual JSONs by expedition number."""
    index = {}
    for f in os.listdir(INDIVIDUAL_DIR):
        if not f.endswith('.json'):
            continue
        with open(os.path.join(INDIVIDUAL_DIR, f)) as fh:
            d = json.load(fh)
        
        # Extract expedition from cross_ref or id
        exp = None
        if '_cross_ref' in d:
            exp = d['_cross_ref'].get('excel_exp')
        if not exp:
            # Try from id
            jid = d.get('id', '')
            m = re.search(r'(\d{1,4})-(\d{4})', jid)
            if m:
                exp = f"{int(m.group(1))}/{m.group(2)}"
        
        if exp:
            index[exp] = d
    
    return index


# ─── MAIN FIX SCRIPT ───────────────────────────────────────────

def main():
    print("=" * 60)
    print("  CIAF Visor Fixer")
    print("  Corrige gravedad, tipología, estaciones y geolocalización")
    print("=" * 60)
    
    # Load data sources
    print("\n📊 Cargando datos...")
    station_db = load_station_coords()
    print(f"  Station coords DB: {len(station_db)} entradas")
    
    individual_index = build_individual_index()
    print(f"  Individual JSONs indexados: {len(individual_index)}")
    
    excel_data = load_excel_data()
    print(f"  Excel records: {len(excel_data)}")
    
    # Process each year file
    stats = {
        'gravedad_fixed': 0,
        'tipo_fixed': 0,
        'station_cleaned': 0,
        'geocoded': 0,
        'total_records': 0,
    }
    
    year_files = sorted([f for f in os.listdir(REPORT_DIR) if f.endswith('.json')])
    
    for yf in year_files:
        year = yf.replace('.json', '')
        fpath = os.path.join(REPORT_DIR, yf)
        
        with open(fpath) as f:
            records = json.load(f)
        
        fixed_records = []
        year_stats = {'gravedad': 0, 'tipo': 0, 'station': 0, 'geo': 0}
        
        for r in records:
            stats['total_records'] += 1
            exp = r.get('expediente', '')
            
            # Get individual JSON data for this expedition
            ind_data = individual_index.get(exp, {})
            
            # Get Excel data
            excel_row = excel_data.get(exp, {})
            
            # 1. Fix gravedad
            old_gravedad = r.get('gravedad', '')
            new_gravedad = GRAVEDAD_MAP.get(old_gravedad, old_gravedad)
            
            # If we have Excel data, use it for precision
            if excel_row:
                muertos = excel_row.get('muertos', 0) or 0
                hg = excel_row.get('heridos_graves', 0) or 0
                if muertos > 0:
                    new_gravedad = "muy grave"
                elif hg > 0:
                    new_gravedad = "grave"
                elif old_gravedad == 'fatal':
                    new_gravedad = "muy grave"  # fallback
            
            if new_gravedad != old_gravedad:
                r['gravedad'] = new_gravedad
                year_stats['gravedad'] += 1
                stats['gravedad_fixed'] += 1
            
            # 2. Fix tipo
            old_tipo = r.get('tipo', '')
            tipo_suceso = r.get('tipo_suceso', '') or excel_row.get('tipo_suceso', '')
            
            if tipo_suceso:
                tipo_lower = tipo_suceso.lower().strip()
                if tipo_lower in TIPO_MAP:
                    cat, detalle = TIPO_MAP[tipo_lower]
                    r['tipo'] = cat
                    r['tipo_suceso'] = tipo_suceso
                    r['tipo_suceso_normalizado'] = detalle
                    if cat != old_tipo:
                        year_stats['tipo'] += 1
                        stats['tipo_fixed'] += 1
            elif ind_data:
                # Use individual JSON data
                if ind_data.get('tipo_suceso'):
                    r['tipo_suceso'] = ind_data['tipo_suceso']
                if ind_data.get('tipo_suceso_normalizado'):
                    r['tipo_suceso_normalizado'] = ind_data['tipo_suceso_normalizado']
            
            # 3. Fix station name
            loc = r.get('ubicacion', {})
            old_station = loc.get('estacion', '')
            if old_station:
                new_station = clean_station_name(old_station)
                if new_station != old_station:
                    loc['estacion'] = new_station
                    year_stats['station'] += 1
                    stats['station_cleaned'] += 1
            
            # 4. Geocode if missing
            if not loc.get('lat') or not loc.get('lng'):
                station_name = loc.get('estacion', '')
                lat, lng = geocode_station(station_name, station_db)
                if lat and lng:
                    loc['lat'] = lat
                    loc['lng'] = lng
                    year_stats['geo'] += 1
                    stats['geocoded'] += 1
            
            # 5. Add detailed fields from individual JSON if available
            if ind_data:
                if not r.get('consecuencias', {}).get('victimas_fallecidos') and ind_data.get('victimas_fallecidos'):
                    r.setdefault('consecuencias', {})
                    r['consecuencias']['victimas_fallecidos'] = ind_data['victimas_fallecidos']
                    r['consecuencias']['victimas_graves'] = ind_data.get('victimas_graves', 0)
                    r['consecuencias']['victimas_leves'] = ind_data.get('victimas_leves', 0)
            
            fixed_records.append(r)
        
        # Save fixed file
        with open(fpath, 'w', encoding='utf-8') as f:
            json.dump(fixed_records, f, ensure_ascii=False, indent=2)
        
        if any(year_stats.values()):
            print(f"  {year}: g={year_stats['gravedad']} t={year_stats['tipo']} s={year_stats['station']} geo={year_stats['geo']}")
    
    # Also update index.json
    index_path = os.path.join(REPORT_DIR, "index.json")
    if os.path.exists(index_path):
        with open(index_path) as f:
            index = json.load(f)
        
        # Update gravedad values in index
        for item in index:
            if isinstance(item, dict):
                g = item.get('gravedad', '')
                if g in GRAVEDAD_MAP:
                    item['gravedad'] = GRAVEDAD_MAP[g]
        
        with open(index_path, 'w', encoding='utf-8') as f:
            json.dump(index, f, ensure_ascii=False, indent=2)
    
    # Final stats
    print(f"\n{'=' * 60}")
    print(f"  RESUMEN")
    print(f"{'=' * 60}")
    print(f"  Registros procesados: {stats['total_records']}")
    print(f"  Severidad corregida: {stats['gravedad_fixed']}")
    print(f"  Tipología corregida: {stats['tipo_fixed']}")
    print(f"  Estaciones limpiadas: {stats['station_cleaned']}")
    print(f"  Geolocalizados: {stats['geocoded']}")
    
    # Verify final distribution
    print(f"\n📊 Distribución final de gravedad:")
    final_gravedad = {}
    for yf in year_files:
        with open(os.path.join(REPORT_DIR, yf)) as f:
            data = json.load(f)
        for r in data:
            g = r.get('gravedad', 'N/A')
            final_gravedad[g] = final_gravedad.get(g, 0) + 1
    
    for k, v in sorted(final_gravedad.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    
    # Count final geocoded
    final_coords = 0
    for yf in year_files:
        with open(os.path.join(REPORT_DIR, yf)) as f:
            data = json.load(f)
        for r in data:
            loc = r.get('ubicacion', {})
            if loc.get('lat') and loc.get('lng'):
                final_coords += 1
    
    print(f"\n📍 Total geolocalizados: {final_coords}/{stats['total_records']}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
